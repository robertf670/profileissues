"""Build a formatted .xlsx audit workbook from the results DataFrame."""

from __future__ import annotations

from io import BytesIO
from typing import TypedDict

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class TripMeta(TypedDict):
    route: str
    direction: str
    terminus_departure: str
    service_date_display: str
    day_type: str


# Five meta rows (1–5), blank row 6, segment header row 7 → pandas startrow=6 (0-indexed).
_TABLE_STARTROW = 6


def build_audit_excel_bytes(df: pd.DataFrame, meta: TripMeta) -> bytes:
    """Trip info once at top; then segment table with headers, widths, formats."""
    buf = BytesIO()
    sheet_name = "Schedule audit"

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=_TABLE_STARTROW)
        ws = writer.sheets[sheet_name]

        label_font = Font(bold=True, size=11)
        meta_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        thin = Side(style="thin", color="CBD5E1")
        grid = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws["A1"], ws["B1"] = "Route number", meta["route"]
        ws["A2"], ws["B2"] = "Direction", meta["direction"]
        ws["A3"], ws["B3"] = "Terminus departure (scheduled)", meta["terminus_departure"]
        ws["A4"], ws["B4"] = "Service date", meta["service_date_display"]
        ws["A5"], ws["B5"] = "Day type", meta["day_type"]
        for r in (1, 2, 3, 4, 5):
            ws.cell(row=r, column=1).font = label_font
            ws.cell(row=r, column=1).fill = meta_fill
            ws.cell(row=r, column=2).fill = meta_fill
            ws.cell(row=r, column=1).border = grid
            ws.cell(row=r, column=2).border = grid
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")

        header_row = _TABLE_STARTROW + 1  # 1-based Excel row of segment column headers
        header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = grid

        fmt_by_col = {
            "Distance along shape (m)": "0.0",
            "Scheduled time (s)": "0",
            "Implied speed (km/h)": "0.00",
        }

        for r in range(header_row + 1, ws.max_row + 1):
            for c_idx, col_name in enumerate(df.columns, start=1):
                cell = ws.cell(row=r, column=c_idx)
                cell.border = grid
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if col_name in fmt_by_col and cell.value is not None and cell.value != "":
                    cell.number_format = fmt_by_col[col_name]

        for c_idx, col_name in enumerate(df.columns, start=1):
            letter = get_column_letter(c_idx)
            sample = [len(str(col_name))]
            if len(df) > 0:
                sample.append(int(df[col_name].astype(str).str.len().max()))
            w = min(max(sample) + 2, 52)
            ws.column_dimensions[letter].width = w

        ws.column_dimensions["A"].width = max(28, ws.column_dimensions["A"].width or 0)
        ws.column_dimensions["B"].width = max(24, ws.column_dimensions["B"].width or 0)

        ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
        ws.sheet_view.zoomScale = 110
        if ws.max_row >= header_row:
            last_col = get_column_letter(len(df.columns))
            ws.auto_filter.ref = f"A{header_row}:{last_col}{ws.max_row}"

    buf.seek(0)
    return buf.getvalue()
